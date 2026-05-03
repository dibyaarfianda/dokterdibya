from __future__ import annotations

import copy
import json
import math
import random
import re
import time
from collections import Counter
from dataclasses import dataclass
from typing import Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Color, PatternFill


@dataclass
class StaffMember:
    no: int
    name: str
    row: int


@dataclass
class SchedulerConfig:
    sheet_name: str = "JADWAL BARU"
    start_day: int = 1
    end_day: int = 31

    max_core_rank: int = 14
    p_count: int = 3
    s_count: int = 3
    m_count: int = 3

    iterations: int = 60000
    initial_temperature: float = 4.5
    cooling_rate: float = 0.99
    cooling_step: int = 3000
    random_seed: int = 707

    tandem_senior_max_rank: int = 11
    top_rank_count: int = 2
    top_rank_max_night: int = 3
    max_consecutive_work: int = 6

    # Example for rank 1..14: 12,12,11,11,10,10,9,9,9,9,9,8,8,8
    off_targets_by_rank: Optional[List[int]] = None

    # Example for rank 12,13,14: [12, 13, 14]
    uniform_group_ranks: Optional[List[int]] = None

    # MAGANG processing controls.
    process_magang: bool = False
    # Example keywords: ["magang", "intern", "koas", "praktek"]
    magang_keywords: Optional[List[str]] = None
    # Optional explicit rank numbers to force as MAGANG participants.
    magang_ranks: Optional[List[int]] = None

    enforce_no_m_to_p: bool = True
    enforce_mll_each: bool = True
    enforce_rank_group_not_together: bool = True
    enforce_tandem: bool = True
    enforce_top_rank_night_cap: bool = True
    enforce_uniform_group_off: bool = True
    enforce_uniform_group_night: bool = True
    enforce_night_monotonic: bool = True
    enforce_off_monotonic: bool = True

    assign_colors: bool = True
    yellow_only_max_rank: int = 11
    polos_per_shift: int = 2
    yellow_per_shift: int = 1
    keep_x_gray: bool = True


class OfflineSchedulerEngine:
    """Generates and validates offline duty schedules in an Excel template."""

    def __init__(self) -> None:
        self.fill_yellow = PatternFill(fill_type="solid", fgColor="FFFFFF00")
        self.fill_plain = PatternFill(fill_type=None)
        self.fill_gray = PatternFill(fill_type="solid", fgColor="FFA6A6A6")
        self.fill_green = PatternFill(fill_type="solid", fgColor="FFE2EFDA")

    @staticmethod
    def _default_magang_keywords() -> List[str]:
        return ["magang", "intern", "koas", "praktek"]

    @staticmethod
    def _normalize_magang_keywords(keywords: Optional[List[str]]) -> List[str]:
        src = keywords if keywords else OfflineSchedulerEngine._default_magang_keywords()
        out: List[str] = []
        seen = set()
        for item in src:
            token = str(item or "").strip().lower()
            if not token or token in seen:
                continue
            seen.add(token)
            out.append(token)
        return out

    @staticmethod
    def _default_magang_duty_target(active_day_count: int) -> int:
        if active_day_count <= 0:
            return 0
        # Baseline from operational note: around 6 duties per 28 active days.
        target = int(round(active_day_count * (6.0 / 28.0)))
        return max(1, min(active_day_count, target))

    def _apply_magang_schedule(
        self,
        schedule: Dict[str, Dict[str, object]],
        magang_staff: List[StaffMember],
        active_days: List[int],
    ) -> None:
        if not magang_staff or not active_days:
            return

        target_duties = self._default_magang_duty_target(len(active_days))
        if target_duties <= 0:
            return

        shift_cycle = ["P", "S", "M"]
        total_days = len(active_days)

        for idx, staff in enumerate(sorted(magang_staff, key=lambda s: s.no)):
            name = staff.name
            if name not in schedule:
                continue

            codes = schedule[name]["codes"]
            for d in active_days:
                codes[d] = "L"

            picked_positions = set()
            for k in range(target_duties):
                base_pos = int(round(((k + 0.5) * total_days / target_duties) - 0.5))
                base_pos = max(0, min(total_days - 1, base_pos))
                pos = (base_pos + idx) % total_days
                guard = 0
                while pos in picked_positions and guard < total_days:
                    pos = (pos + 1) % total_days
                    guard += 1
                picked_positions.add(pos)

            for order, pos in enumerate(sorted(picked_positions)):
                d = active_days[pos]
                codes[d] = shift_cycle[(idx + order) % len(shift_cycle)]

    def _clear_note_rows(self, ws, day_cols: Dict[int, int]) -> None:
        for r in range(3, ws.max_row + 1):
            no_value = ws.cell(r, 1).value
            if self._parse_rank_no(no_value) is not None:
                continue

            note_text = str(ws.cell(r, 2).value or "").strip()
            if not note_text or not self._looks_like_note_line(note_text):
                continue

            for d in sorted(day_cols):
                col = day_cols[d]
                cell = ws.cell(r, col)
                cell.value = None
                cell.fill = self.fill_plain
                self._set_readable_font(cell)

    @staticmethod
    def _is_magang_name(name: str, keywords: Optional[List[str]] = None) -> bool:
        name_l = str(name or "").strip().lower()
        if not name_l:
            return False
        for token in OfflineSchedulerEngine._normalize_magang_keywords(keywords):
            if token in name_l:
                return True
        return False

    def _is_magang_staff(self, staff: StaffMember, config: SchedulerConfig) -> bool:
        if not config.process_magang:
            return False

        rank_set = set()
        for r in config.magang_ranks or []:
            try:
                parsed = int(r)
            except Exception:
                continue
            if parsed > 0:
                rank_set.add(parsed)

        if rank_set and staff.no in rank_set:
            return True

        # Fallback for common template where MAGANG is the immediate rank after core team.
        if not rank_set and staff.no == config.max_core_rank + 1:
            return True

        return self._is_magang_name(staff.name, config.magang_keywords)

    def _is_core_staff(self, staff: StaffMember, config: SchedulerConfig) -> bool:
        magang_by_name = self._is_magang_name(staff.name, config.magang_keywords)

        if config.process_magang:
            return staff.no <= config.max_core_rank or self._is_magang_staff(staff, config)

        # When MAGANG processing is disabled, exclude MAGANG rows from duty generation.
        return staff.no <= config.max_core_rank and not magang_by_name

    @staticmethod
    def _parse_rank_no(value) -> Optional[int]:
        if isinstance(value, bool):
            return None

        if isinstance(value, int):
            return value if value > 0 else None

        if isinstance(value, float):
            if not math.isfinite(value):
                return None
            if float(value).is_integer() and value > 0:
                return int(value)
            return None

        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            m = re.fullmatch(r"(\d+)(?:(?:[\.,]0*)?)", text)
            if not m:
                return None
            try:
                parsed = int(m.group(1))
                return parsed if parsed > 0 else None
            except Exception:
                return None

        return None

    @staticmethod
    def _looks_like_note_line(text: str) -> bool:
        t = str(text or "").strip().lower()
        if not t:
            return False
        if any(ch in t for ch in [":", "|", "="]):
            return True
        if len(t.split()) >= 7:
            return True
        if re.search(r"\b(keterangan|kuning|polos|libur|tandem|hitungan|jadwal terpisah)\b", t):
            return True
        return False

    @staticmethod
    def _set_readable_font(cell) -> None:
        # Preserve existing font shape while forcing readable dark text.
        f = copy.copy(cell.font)
        f.color = Color(rgb="FF000000")
        cell.font = f

    def analyze_workbook(self, input_path: str, config: SchedulerConfig) -> Dict[str, object]:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        ws = wb[config.sheet_name]

        day_cols = self._parse_day_columns(ws)
        staff_all = self._parse_staff(ws, config)
        core_staff = [s for s in staff_all if self._is_core_staff(s, config)]
        team_staff = [s for s in core_staff if not self._is_magang_staff(s, config)]
        core_names = [s.name for s in core_staff]
        team_names = [s.name for s in team_staff]
        row_by_name = {s.name: s.row for s in core_staff}
        no_by_name = {s.name: s.no for s in team_staff}

        active_days = [d for d in sorted(day_cols) if config.start_day <= d <= config.end_day]

        def code_of(name: str, day: int) -> str:
            v = ws.cell(row_by_name[name], day_cols[day]).value
            return self._norm_code(v)

        coverage_bad = []
        tandem_bad = []
        group_together_bad = []
        m_to_p_pairs = []

        expected_l = len(team_staff) - (config.p_count + config.s_count + config.m_count)

        for day in active_days:
            cnt = Counter(code_of(n, day) for n in team_names)
            if not (
                cnt["P"] == config.p_count
                and cnt["S"] == config.s_count
                and cnt["M"] == config.m_count
                and cnt["L"] == expected_l
            ):
                coverage_bad.append(
                    {
                        "day": day,
                        "P": cnt["P"],
                        "S": cnt["S"],
                        "M": cnt["M"],
                        "L": cnt["L"],
                    }
                )

            for sh in ["P", "S", "M"]:
                participants = [n for n in team_names if code_of(n, day) == sh]
                juniors = [n for n in participants if no_by_name[n] > config.tandem_senior_max_rank]
                seniors = [n for n in participants if no_by_name[n] <= config.tandem_senior_max_rank]

                if juniors and not seniors:
                    tandem_bad.append({"day": day, "shift": sh, "participants": participants})

                if config.uniform_group_ranks:
                    in_group = [
                        n for n in participants if no_by_name[n] in set(config.uniform_group_ranks)
                    ]
                    if len(in_group) > 1:
                        group_together_bad.append(
                            {"day": day, "shift": sh, "rank_group": in_group}
                        )

        for n in team_names:
            for day in active_days[:-1]:
                a = code_of(n, day)
                b = code_of(n, day + 1)
                if a == "M" and b == "P":
                    m_to_p_pairs.append({"name": n, "pair": [day, day + 1]})

        off_by_rank = []
        m_by_rank = []
        mll_missing = []

        for s in team_staff:
            codes = {d: code_of(s.name, d) for d in active_days}
            off = sum(1 for d in active_days if codes[d] == "L")
            night = sum(1 for d in active_days if codes[d] == "M")
            mll = self._mll_hits(codes, active_days)

            off_by_rank.append({"rank": s.no, "name": s.name, "L": off})
            m_by_rank.append({"rank": s.no, "name": s.name, "M": night})
            if not mll:
                mll_missing.append(s.name)

        off_monotonic_bad = []
        for i in range(len(off_by_rank) - 1):
            a = off_by_rank[i]
            b = off_by_rank[i + 1]
            if a["L"] < b["L"]:
                off_monotonic_bad.append(
                    {
                        "higher_rank": a["rank"],
                        "higher_name": a["name"],
                        "higher_off": a["L"],
                        "lower_rank": b["rank"],
                        "lower_name": b["name"],
                        "lower_off": b["L"],
                    }
                )

        color_errors = []
        if config.assign_colors:
            for day in active_days:
                col = day_cols[day]
                for sh in ["P", "S", "M"]:
                    participants = [n for n in team_names if code_of(n, day) == sh]
                    yellow_count = 0
                    nofill_count = 0
                    yellow_not_senior = []

                    for n in participants:
                        cell = ws.cell(row_by_name[n], col)
                        rgb = cell.fill.fgColor.rgb
                        if rgb == "FFFFFF00":
                            yellow_count += 1
                            if no_by_name[n] > config.yellow_only_max_rank:
                                yellow_not_senior.append(n)
                        if cell.fill.fill_type is None:
                            nofill_count += 1

                    if (
                        yellow_count != config.yellow_per_shift
                        or nofill_count != config.polos_per_shift
                        or yellow_not_senior
                    ):
                        color_errors.append(
                            {
                                "day": day,
                                "shift": sh,
                                "yellow": yellow_count,
                                "nofill": nofill_count,
                                "yellow_not_rank_1_11": yellow_not_senior,
                            }
                        )

        return {
            "file": input_path,
            "coverage_bad_count": len(coverage_bad),
            "tandem_bad_count": len(tandem_bad),
            "group_together_bad_count": len(group_together_bad),
            "m_to_p_count": len(m_to_p_pairs),
            "mll_missing_count": len(mll_missing),
            "off_monotonic_bad_count": len(off_monotonic_bad),
            "color_errors_count": len(color_errors),
            "off_by_rank": off_by_rank,
            "m_by_rank": m_by_rank,
            "coverage_bad_preview": coverage_bad[:5],
            "off_monotonic_bad_preview": off_monotonic_bad[:5],
            "color_errors_preview": color_errors[:5],
        }

    def generate_schedule(
        self,
        input_path: str,
        output_path: str,
        config: SchedulerConfig,
        progress_callback: Optional[Callable[[Dict[str, object]], None]] = None,
        save_fallback_paths: Optional[List[str]] = None,
    ) -> Dict[str, object]:
        started = time.perf_counter()
        self._emit_progress(
            progress_callback,
            phase="prepare",
            progress=0.0,
            message="Preparing workbook and initial schedule",
            elapsed_sec=0.0,
            eta_sec=None,
            iteration=0,
            total_iterations=config.iterations,
            best_score=None,
            temperature=config.initial_temperature,
        )
        # Yield once so GUI thread can render initial progress text/spinner immediately.
        time.sleep(0)

        random.seed(config.random_seed)

        wb = openpyxl.load_workbook(input_path)
        ws = wb[config.sheet_name]

        day_cols = self._parse_day_columns(ws)
        all_staff = self._parse_staff(ws, config)
        core_staff = [s for s in all_staff if self._is_core_staff(s, config)]
        core_staff = sorted(core_staff, key=lambda s: s.no)
        team_staff = [s for s in core_staff if not self._is_magang_staff(s, config)]
        team_staff = sorted(team_staff, key=lambda s: s.no)
        if not team_staff:
            raise ValueError("No non-MAGANG team rows found for coverage generation")

        core_names = [s.name for s in core_staff]
        team_names = [s.name for s in team_staff]
        row_by_name = {s.name: s.row for s in core_staff}
        no_by_name = {s.name: s.no for s in team_staff}
        name_by_no = {s.no: s.name for s in team_staff}

        active_days = [d for d in sorted(day_cols) if config.start_day <= d <= config.end_day]
        if not active_days:
            raise ValueError("No active days found in selected day range")
        all_days = sorted(day_cols)
        inactive_days = [d for d in all_days if d not in active_days]

        expected_l = len(team_staff) - (config.p_count + config.s_count + config.m_count)
        if expected_l < 0:
            raise ValueError("Coverage is larger than available core staff count")

        # Build initial schedule from file, then rebuild invalid days if needed.
        schedule = {}
        for s in core_staff:
            codes = {}
            for d in active_days:
                v = ws.cell(s.row, day_cols[d]).value
                code = self._norm_code(v)
                if code not in {"P", "S", "M", "L"}:
                    code = "L"
                codes[d] = code
            schedule[s.name] = {"no": s.no, "codes": codes}

        for d in active_days:
            cnt = Counter(schedule[n]["codes"][d] for n in team_names)
            if not (
                cnt["P"] == config.p_count
                and cnt["S"] == config.s_count
                and cnt["M"] == config.m_count
                and cnt["L"] == expected_l
            ):
                self._randomize_day(schedule, team_names, d, config)

        uniform_group = self._resolve_uniform_group(config, team_staff)
        total_off_required = expected_l * len(active_days)
        off_targets = self._resolve_off_targets(
            config,
            team_staff,
            total_off_required=total_off_required,
            uniform_group=uniform_group,
        )

        # Simulated annealing with same-day swaps.
        current = copy.deepcopy(schedule)
        current_score = self._score(
            current,
            team_names,
            name_by_no,
            no_by_name,
            active_days,
            off_targets,
            uniform_group,
            config,
        )
        best = copy.deepcopy(current)
        best_score = current_score

        t = max(config.initial_temperature, 0.01)
        progress_step = max(1, min(1000, max(config.iterations, 1) // 240))

        for it in range(config.iterations):
            day = random.choice(active_days)
            if len(team_names) < 2:
                break
            a, b = random.sample(team_names, 2)
            ca = current[a]["codes"][day]
            cb = current[b]["codes"][day]
            if ca == cb:
                continue

            current[a]["codes"][day], current[b]["codes"][day] = cb, ca

            new_score = self._score(
                current,
                team_names,
                name_by_no,
                no_by_name,
                active_days,
                off_targets,
                uniform_group,
                config,
            )
            delta = new_score - current_score

            if delta >= 0 or random.random() < math.exp(delta / max(t, 0.0001)):
                current_score = new_score
                if new_score > best_score:
                    best_score = new_score
                    best = copy.deepcopy(current)
            else:
                current[a]["codes"][day], current[b]["codes"][day] = ca, cb

            if it > 0 and it % max(config.cooling_step, 1) == 0:
                t *= config.cooling_rate

            if it % progress_step == 0 or it == config.iterations - 1:
                elapsed = time.perf_counter() - started
                current_iter = it + 1
                progress = current_iter / max(config.iterations, 1)
                eta = None
                if progress > 0:
                    eta = max(0.0, (elapsed / progress) - elapsed)

                self._emit_progress(
                    progress_callback,
                    phase="optimize",
                    progress=progress,
                    message="Optimizing schedule",
                    elapsed_sec=elapsed,
                    eta_sec=eta,
                    iteration=current_iter,
                    total_iterations=config.iterations,
                    best_score=best_score,
                    temperature=t,
                )
                # Release GIL periodically so UI updates are not starved by CPU-heavy loops.
                time.sleep(0)

        self._emit_progress(
            progress_callback,
            phase="save",
            progress=0.995,
            message="Writing output file and running final analysis",
            elapsed_sec=time.perf_counter() - started,
            eta_sec=0.0,
            iteration=config.iterations,
            total_iterations=config.iterations,
            best_score=best_score,
            temperature=t,
        )

        best, best_score = self._repair_top_off_targets(
            schedule=best,
            current_score=best_score,
            team_names=team_names,
            name_by_no=name_by_no,
            no_by_name=no_by_name,
            active_days=active_days,
            off_targets=off_targets,
            uniform_group=uniform_group,
            config=config,
        )

        magang_staff = [s for s in core_staff if self._is_magang_staff(s, config)]
        self._apply_magang_schedule(best, magang_staff, active_days)

        # Write optimized codes.
        for n in core_names:
            r = row_by_name[n]
            for d in active_days:
                cell = ws.cell(r, day_cols[d])
                cell.value = best[n]["codes"][d]
                self._set_readable_font(cell)

            # Clear stale highlight on days outside selected generation range.
            for d in inactive_days:
                cell = ws.cell(r, day_cols[d])
                cell.fill = self.fill_plain
                self._set_readable_font(cell)

        # Any row excluded from core processing must remain off-duty.
        core_rows = {s.row for s in core_staff}
        inactive_rows = [s for s in all_staff if s.row not in core_rows]
        for s in inactive_rows:
            for d in all_days:
                cell = ws.cell(s.row, day_cols[d])
                cell.value = "L"
                cell.fill = self.fill_plain
                self._set_readable_font(cell)

        # Keep explanatory note rows clean from stale day values.
        self._clear_note_rows(ws, day_cols)

        # Optional coloring policy.
        if config.assign_colors:
            self._apply_coloring_policy(
                ws,
                day_cols,
                active_days,
                all_staff,
                core_staff,
                team_staff,
                best,
                config,
            )

        save_targets = [str(output_path)]
        if save_fallback_paths:
            for p in save_fallback_paths:
                if not p:
                    continue
                sp = str(p)
                if sp not in save_targets:
                    save_targets.append(sp)

        save_errors: List[str] = []
        actual_output_path: Optional[str] = None

        for idx, target in enumerate(save_targets):
            try:
                if idx > 0:
                    self._emit_progress(
                        progress_callback,
                        phase="save",
                        progress=0.997,
                        message=f"Retry saving output: {target}",
                        elapsed_sec=time.perf_counter() - started,
                        eta_sec=0.0,
                        iteration=config.iterations,
                        total_iterations=config.iterations,
                        best_score=best_score,
                        temperature=t,
                    )
                wb.save(target)
                actual_output_path = target
                break
            except PermissionError as exc:
                save_errors.append(f"{target}: {exc}")

        if actual_output_path is None:
            joined = "\n".join(save_errors)
            raise PermissionError(
                "Failed to save workbook to any output path.\n"
                f"Requested output: {output_path}\n"
                f"Tried paths:\n{joined}"
            )

        report = self.analyze_workbook(actual_output_path, config)
        report["source_file"] = input_path
        report["output_file"] = actual_output_path
        report["_requested_output_file"] = output_path
        report["_save_fallback_used"] = actual_output_path != str(output_path)
        if save_errors:
            report["_save_errors"] = save_errors
        report["solver_score"] = best_score

        self._emit_progress(
            progress_callback,
            phase="done",
            progress=1.0,
            message="Done",
            elapsed_sec=time.perf_counter() - started,
            eta_sec=0.0,
            iteration=config.iterations,
            total_iterations=config.iterations,
            best_score=best_score,
            temperature=t,
        )

        return report

    def _parse_day_columns(self, ws) -> Dict[int, int]:
        out = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            try:
                d = int(v)
                if 1 <= d <= 31 and d not in out:
                    out[d] = c
            except Exception:
                continue
        if not out:
            raise ValueError("Day columns not found in header row 2")
        return out

    def _parse_staff(self, ws, config: Optional[SchedulerConfig] = None) -> List[StaffMember]:
        staff = []
        magang_rows: List[Tuple[int, str]] = []
        max_no = 0
        used_nos = set()

        for r in range(3, ws.max_row + 1):
            name = ws.cell(r, 2).value
            name_clean = str(name or "").strip()
            if not name_clean:
                continue

            no_value = ws.cell(r, 1).value
            parsed_no = self._parse_rank_no(no_value)
            if parsed_no is not None:
                assigned_no = parsed_no
                if assigned_no in used_nos:
                    assigned_no = max_no + 1

                staff.append(StaffMember(no=assigned_no, name=name_clean, row=r))
                used_nos.add(assigned_no)
                if assigned_no > max_no:
                    max_no = assigned_no
                continue

            keywords = config.magang_keywords if config else None
            if self._is_magang_name(name_clean, keywords):
                if self._looks_like_note_line(name_clean):
                    continue
                if config is not None and not config.process_magang:
                    continue
                magang_rows.append((r, name_clean))

        for r, name_clean in magang_rows:
            next_no = max_no + 1
            while next_no in used_nos:
                next_no += 1
            max_no = next_no
            staff.append(StaffMember(no=max_no, name=name_clean, row=r))
            used_nos.add(max_no)

        if not staff:
            raise ValueError("No staff rows found (expects NO in col A and NAMA in col B)")
        return sorted(staff, key=lambda s: s.no)

    def _norm_code(self, value) -> str:
        code = str(value).strip().upper() if value is not None else ""
        if code == "O":
            return "L"
        return code

    def _randomize_day(
        self,
        schedule: Dict[str, Dict[str, object]],
        core_names: List[str],
        day: int,
        config: SchedulerConfig,
    ) -> None:
        names = core_names[:]
        random.shuffle(names)

        p_slice = names[: config.p_count]
        s_slice = names[config.p_count : config.p_count + config.s_count]
        m_slice = names[
            config.p_count + config.s_count : config.p_count + config.s_count + config.m_count
        ]

        for n in core_names:
            schedule[n]["codes"][day] = "L"
        for n in p_slice:
            schedule[n]["codes"][day] = "P"
        for n in s_slice:
            schedule[n]["codes"][day] = "S"
        for n in m_slice:
            schedule[n]["codes"][day] = "M"

    def _resolve_off_targets(
        self,
        config: SchedulerConfig,
        core_staff: List[StaffMember],
        total_off_required: int,
        uniform_group: List[int],
    ) -> Dict[str, int]:
        ordered_staff = sorted(core_staff, key=lambda s: s.no)

        if config.off_targets_by_rank:
            base_values = []
            for s in ordered_staff:
                idx = s.no - 1
                if 0 <= idx < len(config.off_targets_by_rank):
                    base_values.append(int(config.off_targets_by_rank[idx]))
                else:
                    base_values.append(int(config.off_targets_by_rank[-1]))
        else:
            # Default profile used in the recent schedule iterations for 27 active days.
            default_14 = [12, 12, 11, 11, 10, 10, 9, 9, 9, 9, 9, 8, 8, 8]
            base_values = []
            for s in ordered_staff:
                idx = s.no - 1
                if 0 <= idx < len(default_14):
                    base_values.append(int(default_14[idx]))
                else:
                    base_values.append(int(default_14[-1]))

        adjusted_values = self._rebalance_off_targets(
            ordered_staff=ordered_staff,
            base_values=base_values,
            total_off_required=total_off_required,
            uniform_group=uniform_group,
        )

        return {staff.name: adjusted_values[i] for i, staff in enumerate(ordered_staff)}

    def _rebalance_off_targets(
        self,
        ordered_staff: List[StaffMember],
        base_values: List[int],
        total_off_required: int,
        uniform_group: List[int],
    ) -> List[int]:
        targets = [max(0, int(v)) for v in base_values]
        if not targets:
            return targets

        group_set = set(uniform_group)
        non_group_indices = [i for i, s in enumerate(ordered_staff) if s.no not in group_set]
        group_indices = [i for i, s in enumerate(ordered_staff) if s.no in group_set]

        add_priority = list(reversed(non_group_indices)) if non_group_indices else list(reversed(range(len(targets))))
        remove_priority = list(reversed(non_group_indices)) if non_group_indices else list(reversed(range(len(targets))))
        add_priority_non_top = [idx for idx in add_priority if idx != 0]
        if not add_priority_non_top:
            add_priority_non_top = add_priority[:]

        def _add_with_monotonic(need: int) -> int:
            guard = 0
            while need > 0 and guard < 200000:
                progressed = False
                for idx in add_priority_non_top:
                    prev_limit = targets[idx - 1] if idx > 0 else None
                    if prev_limit is None or targets[idx] < prev_limit:
                        targets[idx] += 1
                        need -= 1
                        progressed = True
                        if need <= 0:
                            break

                if not progressed:
                    # Only lift top rank as a last resort when lower ranks cannot absorb more.
                    if 0 in add_priority:
                        targets[0] += 1
                        need -= 1
                    else:
                        break

                guard += 1
            return need

        def _remove(need: int) -> int:
            guard = 0
            while need > 0 and remove_priority and guard < 200000:
                progressed = False
                for idx in remove_priority:
                    if targets[idx] > 0:
                        targets[idx] -= 1
                        need -= 1
                        progressed = True
                        if need <= 0:
                            break
                if not progressed:
                    break
                guard += 1
            return need

        diff = int(total_off_required - sum(targets))
        if diff > 0:
            _add_with_monotonic(diff)
        elif diff < 0:
            need = -diff
            need = _remove(need)

            # If still needed and uniform group exists, reduce them evenly.
            while need > 0 and group_indices and all(targets[idx] > 0 for idx in group_indices):
                for idx in group_indices:
                    if need <= 0:
                        break
                    targets[idx] -= 1
                    need -= 1

        # Ensure group equality by aligning all group ranks to the minimum in group.
        if group_indices:
            group_min = min(targets[idx] for idx in group_indices)
            for idx in group_indices:
                targets[idx] = group_min

        # Enforce monotonic seniority: higher rank index should not have fewer off-days.
        for i in range(1, len(targets)):
            if targets[i - 1] < targets[i]:
                targets[i] = targets[i - 1]

        # Rebalance any remaining delta introduced by clamping.
        remain = int(total_off_required - sum(targets))
        if remain > 0:
            _add_with_monotonic(remain)
        elif remain < 0:
            _remove(-remain)

        return targets

    def _resolve_uniform_group(
        self,
        config: SchedulerConfig,
        core_staff: List[StaffMember],
    ) -> List[int]:
        all_ranks = {s.no for s in core_staff}
        if config.uniform_group_ranks:
            return [r for r in config.uniform_group_ranks if r in all_ranks]
        return [r for r in [12, 13, 14] if r in all_ranks]

    def _mll_hits(self, codes: Dict[int, str], active_days: List[int]) -> List[int]:
        hits = []
        for day in active_days:
            if day + 2 not in codes:
                continue
            if codes[day] == "M" and codes[day + 1] == "L" and codes[day + 2] == "L":
                hits.append(day)
        return hits

    def _score(
        self,
        schedule: Dict[str, Dict[str, object]],
        team_names: List[str],
        name_by_no: Dict[int, str],
        no_by_name: Dict[str, int],
        active_days: List[int],
        off_targets: Dict[str, int],
        uniform_group: List[int],
        config: SchedulerConfig,
    ) -> float:
        expected_l = len(team_names) - (config.p_count + config.s_count + config.m_count)

        # Hard counters.
        coverage_bad = 0
        tandem_bad = 0
        group_together_bad = 0
        m_to_p_bad = 0

        off = {n: 0 for n in team_names}
        nights = {n: 0 for n in team_names}

        for day in active_days:
            cnt = Counter(schedule[n]["codes"][day] for n in team_names)
            if not (
                cnt["P"] == config.p_count
                and cnt["S"] == config.s_count
                and cnt["M"] == config.m_count
                and cnt["L"] == expected_l
            ):
                coverage_bad += 1

            for sh in ["P", "S", "M"]:
                participants = [n for n in team_names if schedule[n]["codes"][day] == sh]
                if config.enforce_tandem:
                    juniors = [n for n in participants if no_by_name[n] > config.tandem_senior_max_rank]
                    seniors = [n for n in participants if no_by_name[n] <= config.tandem_senior_max_rank]
                    if juniors and not seniors:
                        tandem_bad += 1

                if config.enforce_rank_group_not_together and uniform_group:
                    in_group = [n for n in participants if no_by_name[n] in set(uniform_group)]
                    if len(in_group) > 1:
                        group_together_bad += 1

            for n in team_names:
                code = schedule[n]["codes"][day]
                if code == "L":
                    off[n] += 1
                if code == "M":
                    nights[n] += 1

        if config.enforce_no_m_to_p:
            for n in team_names:
                for day in active_days[:-1]:
                    if (
                        schedule[n]["codes"][day] == "M"
                        and schedule[n]["codes"][day + 1] == "P"
                    ):
                        m_to_p_bad += 1

        # Soft/medium penalties.
        penalties = 0.0
        penalties += coverage_bad * 180000
        penalties += tandem_bad * 180000
        penalties += group_together_bad * 220000
        penalties += m_to_p_bad * 140000

        if config.enforce_top_rank_night_cap:
            top_ranks = [r for r in range(1, config.top_rank_count + 1) if r in name_by_no]
            top_over = 0
            for r in top_ranks:
                n = name_by_no[r]
                top_over += max(0, nights[n] - config.top_rank_max_night)
            penalties += top_over * 140000

        if config.enforce_off_monotonic:
            off_mono_bad = 0
            off_mono_mag = 0
            for rank in range(1, len(team_names)):
                if rank not in name_by_no or rank + 1 not in name_by_no:
                    continue
                a = name_by_no[rank]
                b = name_by_no[rank + 1]
                if off[a] < off[b]:
                    off_mono_bad += 1
                    off_mono_mag += off[b] - off[a]
            penalties += off_mono_bad * 260000
            penalties += off_mono_mag * 40000

        if config.enforce_night_monotonic:
            night_mono_bad = 0
            night_mono_mag = 0
            for rank in range(1, len(team_names)):
                if rank not in name_by_no or rank + 1 not in name_by_no:
                    continue
                a = name_by_no[rank]
                b = name_by_no[rank + 1]
                if nights[a] > nights[b]:
                    night_mono_bad += 1
                    night_mono_mag += nights[a] - nights[b]
            penalties += night_mono_bad * 60000
            penalties += night_mono_mag * 9000

        if config.enforce_uniform_group_off and uniform_group:
            group_off = [off[name_by_no[r]] for r in uniform_group if r in name_by_no]
            if group_off:
                penalties += (max(group_off) - min(group_off)) * 240000

        if config.enforce_uniform_group_night and uniform_group:
            group_night = [nights[name_by_no[r]] for r in uniform_group if r in name_by_no]
            if group_night:
                penalties += (max(group_night) - min(group_night)) * 130000

        # Off target proximity.
        off_target_pen = 0
        for n in team_names:
            rank = no_by_name.get(n, config.max_core_rank)
            rank_weight = 1.0 + max(0.0, (config.max_core_rank - rank) * 0.3)
            off_target_pen += abs(off[n] - off_targets.get(n, off[n])) * rank_weight
        penalties += off_target_pen * 60000

        # Strongly discourage top ranks from exceeding their off-day targets.
        if config.top_rank_count > 0:
            over_pen = 0
            for r in range(1, config.top_rank_count + 1):
                if r not in name_by_no:
                    continue
                n = name_by_no[r]
                target = off_targets.get(n, off[n])
                over = max(0, off[n] - target)
                if over <= 0:
                    continue
                rank_boost = 2.0 if r == 1 else 1.0
                over_pen += over * rank_boost
            penalties += over_pen * 2500000

        if config.enforce_mll_each:
            missing = 0
            for n in team_names:
                hits = self._mll_hits(schedule[n]["codes"], active_days)
                if not hits:
                    missing += 1
            penalties += missing * 50000

        # Optional max consecutive work soft penalty.
        streak_over = 0
        for n in team_names:
            max_streak = 0
            cur = 0
            for day in active_days:
                if schedule[n]["codes"][day] in {"P", "S", "M"}:
                    cur += 1
                    max_streak = max(max_streak, cur)
                else:
                    cur = 0
            streak_over += max(0, max_streak - config.max_consecutive_work)
        penalties += streak_over * 500

        return -penalties

    def _repair_top_off_targets(
        self,
        schedule: Dict[str, Dict[str, object]],
        current_score: float,
        team_names: List[str],
        name_by_no: Dict[int, str],
        no_by_name: Dict[str, int],
        active_days: List[int],
        off_targets: Dict[str, int],
        uniform_group: List[int],
        config: SchedulerConfig,
    ) -> Tuple[Dict[str, Dict[str, object]], float]:
        if config.top_rank_count <= 0:
            return schedule, current_score

        repaired = copy.deepcopy(schedule)

        off = {
            n: sum(1 for d in active_days if repaired[n]["codes"][d] == "L")
            for n in team_names
        }

        top_ranks = [r for r in range(1, config.top_rank_count + 1) if r in name_by_no]
        donor_order = sorted(team_names, key=lambda x: no_by_name[x], reverse=True)

        guard = 0
        while guard < 800:
            guard += 1
            moved = False

            for r in top_ranks:
                top_name = name_by_no[r]
                target = off_targets.get(top_name, off[top_name])
                if off[top_name] <= target:
                    continue

                best_move = None
                best_move_score = current_score

                for day in active_days:
                    if repaired[top_name]["codes"][day] != "L":
                        continue

                    for donor in donor_order:
                        if donor == top_name:
                            continue

                        donor_code = repaired[donor]["codes"][day]
                        if donor_code not in {"P", "S", "M"}:
                            continue

                        repaired[top_name]["codes"][day] = donor_code
                        repaired[donor]["codes"][day] = "L"

                        candidate_score = self._score(
                            repaired,
                            team_names,
                            name_by_no,
                            no_by_name,
                            active_days,
                            off_targets,
                            uniform_group,
                            config,
                        )

                        repaired[top_name]["codes"][day] = "L"
                        repaired[donor]["codes"][day] = donor_code

                        if candidate_score > best_move_score:
                            best_move_score = candidate_score
                            best_move = (day, donor, donor_code)

                    if best_move is not None:
                        break

                if best_move is None:
                    continue

                day, donor, donor_code = best_move
                repaired[top_name]["codes"][day] = donor_code
                repaired[donor]["codes"][day] = "L"
                off[top_name] -= 1
                off[donor] += 1
                current_score = best_move_score
                moved = True

            if not moved:
                break

        return repaired, current_score

    def _apply_coloring_policy(
        self,
        ws,
        day_cols: Dict[int, int],
        active_days: List[int],
        all_staff: List[StaffMember],
        core_staff: List[StaffMember],
        team_staff: List[StaffMember],
        best_schedule: Dict[str, Dict[str, object]],
        config: SchedulerConfig,
    ) -> None:
        core_names = [s.name for s in core_staff]
        team_names = [s.name for s in team_staff]
        team_set = set(team_names)
        non_team_names = [n for n in core_names if n not in team_set]
        row_by_name = {s.name: s.row for s in core_staff}
        no_by_name = {s.name: s.no for s in team_staff}

        # Keep X as gray on days 1-4.
        if config.keep_x_gray:
            for n in core_names:
                r = row_by_name[n]
                for day in range(1, 5):
                    if day not in day_cols:
                        continue
                    col = day_cols[day]
                    code = self._norm_code(ws.cell(r, col).value)
                    if code == "X":
                        ws.cell(r, col).fill = self.fill_gray

        # Active day color assignment.
        for day in active_days:
            col = day_cols[day]

            # L is always plain/no-fill.
            for n in core_names:
                if best_schedule[n]["codes"][day] == "L":
                    ws.cell(row_by_name[n], col).fill = self.fill_plain

            # MAGANG rows are not part of team coloring.
            for n in non_team_names:
                ws.cell(row_by_name[n], col).fill = self.fill_plain

            for sh in ["P", "S", "M"]:
                participants = [n for n in team_names if best_schedule[n]["codes"][day] == sh]
                if not participants:
                    continue

                seniors = [n for n in participants if no_by_name[n] <= config.yellow_only_max_rank]
                if seniors:
                    yellow = sorted(seniors, key=lambda x: no_by_name[x])[0]
                else:
                    yellow = sorted(participants, key=lambda x: no_by_name[x])[0]

                polos = [n for n in participants if n != yellow]

                ws.cell(row_by_name[yellow], col).fill = self.fill_yellow
                for n in polos:
                    ws.cell(row_by_name[n], col).fill = self.fill_plain

        # Keep every row excluded from core generation visually off-duty.
        core_rows = {s.row for s in core_staff}
        inactive_rows = [s for s in all_staff if s.row not in core_rows]
        all_days = sorted(day_cols)
        for s in inactive_rows:
            for day in all_days:
                col = day_cols[day]
                cell = ws.cell(s.row, col)
                cell.value = "L"
                cell.fill = self.fill_plain
                self._set_readable_font(cell)

    @staticmethod
    def _emit_progress(
        progress_callback: Optional[Callable[[Dict[str, object]], None]],
        **payload: object,
    ) -> None:
        if not progress_callback:
            return
        try:
            progress_callback(payload)
        except Exception:
            # Progress updates must never interrupt optimization.
            pass

    @staticmethod
    def report_to_pretty_json(report: Dict[str, object]) -> str:
        return json.dumps(report, indent=2, ensure_ascii=True)
