import json
import openpyxl

path = 'C:/Users/nanda/Desktop/Jadwal VK Mei Revisi.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

out = {'sheets': []}

for ws in wb.worksheets:
    min_r = None
    max_r = None
    min_c = None
    max_c = None
    nonempty = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and v.strip() == '':
                continue
            nonempty += 1
            r, c = cell.row, cell.column
            min_r = r if min_r is None else min(min_r, r)
            max_r = r if max_r is None else max(max_r, r)
            min_c = c if min_c is None else min(min_c, c)
            max_c = c if max_c is None else max(max_c, c)

    sheet_info = {
        'title': ws.title,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'nonempty_cells': nonempty,
        'used_range': None,
        'preview': []
    }

    if min_r is not None:
        sheet_info['used_range'] = {
            'min_row': min_r,
            'max_row': max_r,
            'min_col': min_c,
            'max_col': max_c
        }

        pr_end_r = min(max_r, min_r + 40)
        pr_end_c = min(max_c, min_c + 25)

        for r in range(min_r, pr_end_r + 1):
            vals = []
            has_any = False
            for c in range(min_c, pr_end_c + 1):
                v = ws.cell(r, c).value
                if v is None:
                    vals.append('')
                else:
                    s = str(v).strip()
                    vals.append(s)
                    if s != '':
                        has_any = True
            if has_any:
                sheet_info['preview'].append({'row': r, 'values': vals})

    out['sheets'].append(sheet_info)

print(json.dumps(out, ensure_ascii=False))
